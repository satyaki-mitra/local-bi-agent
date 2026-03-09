# DEPENDENCIES
import io
import csv
import json
import openpyxl
import structlog
import pandas as pd
from typing import Any
from typing import List
from typing import Dict
from typing import Optional
from datetime import datetime
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from config.constants import RESULT_BRANDING_NAME
from config.constants import RESULT_BRANDING_TAGLINE


# Setup Logging
logger = structlog.get_logger()


# Brand palette shared with chart theme (hex without #)
_BRAND_TEAL_HEX   = "00D4AA"
_BRAND_DARK_HEX   = "0D1117"
_BRAND_SURFACE_HEX= "161B22"
_BRAND_TEXT_HEX   = "E6EDF3"
_BRAND_MUTED_HEX  = "8B949E"
_BRAND_BORDER_HEX = "21262D"


class ResultGenerator:
    """
    Generate downloadable outputs in various formats: pure rendering functions — no file I/O.
    Primary formats: CSV, JSON, XLSX. PNG is handled by visualization_generator
    """
    # CSV exports
    def generate_csv_from_dataframe(self, df: pd.DataFrame) -> str:
        """
        Generate CSV from a pandas DataFrame
        """
        output = io.StringIO()
        df.to_csv(output, index = False)
        return output.getvalue()


    def generate_csv_from_data(self, data: List[Dict[str, Any]]) -> str:
        """
        Generate CSV from a list of dictionaries
        """
        if not data:
            return ""

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()

  
    def generate_json(self, query: str, answer: str, data: Optional[List[Dict[str, Any]]] = None, sql_queries: Optional[List[str]] = None,
                      reasoning_trace: Optional[List[str]] = None, metadata: Optional[Dict[str, Any]] = None, analysis: Optional[Dict[str, Any]] = None) -> str:
        """
        Generate a JSON object containing the query result and optional analysis
        """
        result = {"timestamp"       : datetime.now().isoformat(),
                  "query"           : query,
                  "answer"          : answer,
                  "sql_queries"     : sql_queries or [],
                  "reasoning_trace" : reasoning_trace or [],
                  "data"            : data or [],
                  "metadata"        : metadata or {},
                 }

        if analysis:
            # Include the analysis dictionary as a top-level field
            result["analysis"] = analysis

        return json.dumps(result, indent=4, default=str)


    def generate_xlsx(self, data: List[Dict[str, Any]], query: str = "", answer: str = "", sql_queries: Optional[List[str]] = None, analysis: Optional[Dict[str, Any]] = None) -> bytes:
        """
        Generate a styled Excel workbook with up to three sheets:
          - Data (primary result table)
          - Metadata (query, answer, SQL)
          - Analysis (if provided and contains no error)
        """
        wb            = openpyxl.Workbook()

        # Sheet 1: Data
        ws_data       = wb.active
        ws_data.title = "Data"

        if data:
            df           = pd.DataFrame(data)
            headers      = list(df.columns)

            # Styles
            header_font  = Font(name="Calibri", bold=True, size=11, color=_BRAND_DARK_HEX)
            header_fill  = PatternFill(start_color=_BRAND_TEAL_HEX, end_color=_BRAND_TEAL_HEX, fill_type="solid")
            header_align = Alignment(horizontal="left", vertical="center")
            alt_fill     = PatternFill(start_color="1C2128", end_color="1C2128", fill_type="solid")
            body_font    = Font(name="Consolas", size=10, color=_BRAND_TEXT_HEX)
            body_align   = Alignment(horizontal="left", vertical="center", wrap_text=False)
            thin_side    = Side(style="thin", color=_BRAND_BORDER_HEX)
            thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            # Write header row
            for col_idx, header in enumerate(headers, 1):
                cell           = ws_data.cell(row=1, column=col_idx, value=header)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align
                cell.border    = thin_border

            ws_data.row_dimensions[1].height = 22

            # Write data rows
            for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
                use_alt = (row_idx % 2 == 0)
                
                for col_idx, value in enumerate(row_data, 1):
                    cell           = ws_data.cell(row    = row_idx,
                                                  column = col_idx,
                                                  value  = str(value) if value is not None else "",
                                                 )
                    cell.font      = body_font
                    cell.alignment = body_align
                    cell.border    = thin_border
                    
                    if use_alt:
                        cell.fill = alt_fill

            # Auto-fit column widths (capped at 45)
            for col_idx, header in enumerate(headers, 1):
                col_values                                  = [str(header)] + [str(v) for v in df.iloc[:, col_idx - 1].fillna("").head(200)]
                max_len                                     = min(45, max((len(v) for v in col_values), default = 10) + 2)
                col_letter                                  = get_column_letter(col_idx)
                ws_data.column_dimensions[col_letter].width = max_len

            ws_data.freeze_panes = "A2"

        # Sheet 2: Metadata 
        ws_meta       = wb.create_sheet(title="Metadata")
        meta_header_f = Font(name="Calibri", bold=True, size=11, color=_BRAND_TEAL_HEX)
        meta_body_f   = Font(name="Calibri", size=10, color=_BRAND_TEXT_HEX)
        wrap_align    = Alignment(horizontal="left", vertical="top", wrap_text=True)

        meta_rows     = [("Field", "Value"),
                         ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                         ("Query", query),
                         ("Answer", answer[:2000] if answer else ""),
                         ("SQL", "\n".join(sql_queries) if sql_queries else ""),
                         ("Rows", str(len(data))),
                        ]

        for r_idx, (field, value) in enumerate(meta_rows, 1):
            # Field column
            cell_field         = ws_meta.cell(row    = r_idx, 
                                              column = 1, 
                                              value  = field,
                                             )

            cell_field.font    = meta_header_f if r_idx == 1 else Font(name = "Calibri", bold = True, size = 10, color = _BRAND_MUTED_HEX)
            # Value column
            cell_val           = ws_meta.cell(row=r_idx, column=2, value=value)
            cell_val.font      = meta_body_f
            cell_val.alignment = wrap_align

            # Row height
            ws_meta.row_dimensions[r_idx].height = 60 if field in ("Answer", "SQL") else 18

        ws_meta.column_dimensions["A"].width = 14
        ws_meta.column_dimensions["B"].width = 80

        # Sheet 3: Analysis (if provided and meaningful)
        if analysis and isinstance(analysis, dict) and "error" not in analysis:
            ws_analysis = wb.create_sheet(title = "Analysis")
            row         = 1

            # Simple recursive flattening for readability
            def write_dict(d, parent_row, indent=0):
                nonlocal row
                for k, v in d.items():
                    if isinstance(v, dict):
                        ws_analysis.cell(row    = row, 
                                         column = indent+1, 
                                         value  = f"{k}:",
                                        ).font = Font(bold=True)
                        row += 1
                        write_dict(v, row, indent+1)
                    
                    else:
                        ws_analysis.cell(row    = row, 
                                         column = indent+1, 
                                         value  = f"{k}:",
                                        )

                        ws_analysis.cell(row    = row,
                                         column = indent+2, 
                                         value  = str(v)[:200],
                                        )
                        row += 1
            write_dict(analysis, row)

        # Serialise to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()


# GLOBAL INSTANCE
result_generator = ResultGenerator()